import xlsxwriter
import openpyxl
from openpyxl.styles import Alignment

data =[
    {
        'co':'00000002',
        '2023-03-01':'11:06:23',
        '2023-03-01(1)':'13:16:23',
        '2023-02-01':'10:14:11',
	    '2023-02-01(1)':'20:45:30',
        '2023-02-06':'10:14:11',
	    '2023-02-06(1)':'20:45:30',
        '2023-02-08':'10:14:11',
	    '2023-02-08(1)':'20:45:30',
    },
    {
        'co':'00000003',
        '2023-03-01':'11:06:23',
        '2023-03-01(1)':'13:16:23',
        '2023-02-01':'10:14:11',
	    '2023-02-01(1)':'20:45:30',
        '2023-02-06':'10:14:11',
	    '2023-02-06(1)':'20:45:30',
        '2023-02-08':'10:14:11',
	    '2023-02-08(1)':'20:45:30',
    },
    {
        'co':'00000004',
        '2023-03-01':'11:06:23',
        '2023-03-01(1)':'13:16:23',
        '2023-02-01':'10:14:11',
	    '2023-02-01(1)':'20:45:30',
        '2023-02-06':'10:14:11',
	    '2023-02-06(1)':'20:45:30',
        '2023-02-08':'10:14:11',
	    '2023-02-08(1)':'20:45:30',
    }
]

workbook = xlsxwriter.Workbook("workbook.xlsx")
worksheet = workbook.add_worksheet("first")

worksheet.write(0,0,"co")
worksheet.write(0,1,"2023-03-01")
worksheet.write(0,2,"2023-03-01")
worksheet.write(0,3,"2023-02-01")
worksheet.write(0,4,"2023-02-01")
worksheet.write(0,5,"2023-02-06")
worksheet.write(0,6,"2023-02-06")
worksheet.write(0,7,"2023-02-08")
worksheet.write(0,8,"2023-02-08")

for index, entry in enumerate(data):
    worksheet.write(index+1, 0, entry["co"])
    worksheet.write(index+1, 1, entry["2023-03-01"])
    worksheet.write(index+1, 2, entry["2023-03-01(1)"])
    worksheet.write(index+1, 3, entry["2023-02-01"])
    worksheet.write(index+1, 4, entry["2023-02-01(1)"])
    worksheet.write(index+1, 5, entry["2023-02-06"])
    worksheet.write(index+1, 6, entry["2023-02-06(1)"])
    worksheet.write(index+1, 7, entry["2023-02-08"])
    worksheet.write(index+1, 8, entry["2023-02-08(1)"])

workbook.close()
wb = openpyxl.load_workbook('workbook.xlsx')
ws = wb.active
ws.merge_cells('B1:C1')
ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('D1:E1')
ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('F1:G1')
ws['F1'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('H1:I1')
ws['H1'].alignment = Alignment(horizontal='center', vertical='center')
wb.save('workbook.xlsx')